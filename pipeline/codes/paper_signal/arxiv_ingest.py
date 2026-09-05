# -*- coding: utf-8 -*-
"""arXiv 全库批次入库：元数据裸格式 TXT → 仓库标准头块格式，并入 data/papers。

背景（2026-08-27 首批：arxiv_txt_2022.zip）：全库分批交付的论文 TXT 为 arXiv 元数据
裸格式（Title:/Authors:/Published:/Abstract:，仅元数据+摘要，无全文），与 data/papers
现行"竞赛头块 + 全文"格式（paper_parser 解析绑定）不同。本脚本以批次索引 xlsx
（分档/总得分/命中维度/证据句/PDF 直链）为权威字段源，把每个 S/A 档 TXT 转换为
标准头块格式后写入 data/papers/<batch>/<档位目录>/，原始元数据块原样保留为正文；
B/C 档不入库（与全库六专题"仅保留 S/A 档"口径一致，scan_papers 亦只扫 S/A/B）。

用法（项目根目录）：
  python codes/paper_signal/arxiv_ingest.py \
      --src "D:/CodeLib/challenge/temp/arxiv2022/arxiv2022_classify" \
      --batch arxiv2022

入库后重建论文时间线映射表：
  cd codes/timeline && python run_timeline.py --papers

依赖 openpyxl（读批次索引 xlsx）。转换纯确定性、幂等：重复执行覆盖同名产物。
"""
import argparse
import glob
import os
import re
import shutil
import sys

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)
import paper_config as config  # noqa: E402

TIER_DIRNAME = {"S": "S档_核心", "A": "A档_重点"}   # 与既有档位目录命名一致
SEP_LINE = "═" * 51
RAW_DATE_RE = re.compile(r"^Published:\s*(\d{4}-\d{2}-\d{2})")
RAW_TITLE_RE = re.compile(r"^Title:\s*(.+)$", re.M)
EVIDENCE_SPLIT_RE = re.compile(r"\[([^\]]+)\]")


def _find_index_xlsx(src):
    """定位批次索引 xlsx：优先 01_ 前缀（全库总索引），回退 src 根下任一 xlsx。"""
    for pat in ("01_*.xlsx", "*.xlsx"):
        hits = [p for p in glob.glob(os.path.join(src, pat))]
        if hits:
            return hits[0]
    raise FileNotFoundError(f"批次目录下未找到索引 xlsx: {src}")


def load_batch_index(src):
    """读批次索引 xlsx → {arxiv_id: {列名: 值}}。"""
    if openpyxl is None:
        raise ImportError("需要 openpyxl 读取批次索引 xlsx（pip install openpyxl）")
    path = _find_index_xlsx(src)
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() for h in next(rows)]
    index = {}
    for r in rows:
        d = dict(zip(header, r))
        pid = str(d.get("arXiv ID") or "").strip()
        if pid:
            index[pid] = d
    wb.close()
    return index, os.path.basename(path)


def _split_dimensions(cell):
    """'多源数据与新兴岗位发现;简历解析与能力推断' → 维度列表（分号分隔）。"""
    return [d.strip() for d in str(cell or "").split(";") if d.strip()]


def _split_evidence(cell):
    """'[维度1] 证据1 [维度2] 证据2' → [(维度, 证据)]。"""
    text = str(cell or "").strip()
    if not text:
        return []
    marks = list(EVIDENCE_SPLIT_RE.finditer(text))
    out = []
    for i, m in enumerate(marks):
        seg = text[m.end(): marks[i + 1].start() if i + 1 < len(marks) else len(text)]
        seg = re.sub(r"\s+", " ", seg).strip()
        out.append((m.group(1).strip(), seg))
    return [(d, s) for d, s in out if s]


def _date_from_raw(raw_body):
    m = RAW_DATE_RE.search(raw_body)
    return m.group(1) if m else ""


def convert_text(raw_body, meta, tier, batch):
    """原始元数据正文 + 索引行 → 标准头块格式全文。"""
    pid = str(meta.get("arXiv ID") or "").strip()
    m_title = RAW_TITLE_RE.search(raw_body)
    title = str(meta.get("标题") or "").strip() or (m_title.group(1).strip() if m_title else "")
    pub = str(meta.get("发表日期") or "").strip()[:10] or _date_from_raw(raw_body)
    score = meta.get("总得分")
    dims = _split_dimensions(meta.get("命中维度"))
    evidence = _split_evidence(meta.get("证据句"))
    pdf = str(meta.get("PDF直链") or "").strip() or f"https://arxiv.org/pdf/{pid}"
    web = str(meta.get("网页版") or "").strip() or f"http://arxiv.org/abs/{pid}"

    tier_line = f"{tier} 档"
    if isinstance(score, (int, float)) and score:
        tier_line += f"（得分: {int(score)}" + (f", 覆盖 {len(dims)} 个维度" if dims else "") + "）"

    lines = [SEP_LINE, f"【arXiv ID】  {pid}", f"【标题】      {title}",
             f"【发表日期】  {pub}", f"【赛题分档】  {tier_line}"]
    if dims:
        lines.append(f"【命中维度】  {'、'.join(dims)}")
    lines.append(f"【PDF直链】   {pdf}")
    lines.append(f"【网页版】    {web}")
    if evidence:
        lines.append("【证据句】")
        lines.extend(f"  [{d}] {s}" for d, s in evidence)
    lines.append(f"【说明】      本篇为 {batch} 全库批次条目：仅含 arXiv 元数据与摘要，无全文正文")
    lines += [SEP_LINE, config.HEADER_MARKER, ""]
    return "\n".join(lines) + raw_body.strip() + "\n"


def ingest(src, batch="arxiv2022", tiers=("S", "A"), papers_dir=None):
    """转换并写入 data/papers/<batch>/<档位目录>/。返回统计 dict。"""
    papers_dir = papers_dir or config.PAPER_DIR
    dest_root = os.path.join(papers_dir, batch)
    index, index_name = load_batch_index(src)

    # 跨批次同 ID 防线：现有语料文件名 = arXiv ID.txt（scan_papers 按文件名+md5 去重，
    # 内容不同不会去重 → 映射表会出现重复 ID 行），入库前显式警告
    existing_ids = set()
    for d in os.listdir(papers_dir):
        sub = os.path.join(papers_dir, d)
        if not os.path.isdir(sub):
            continue
        for td in os.listdir(sub):
            tdir = os.path.join(sub, td)
            if os.path.isdir(tdir) and re.match(r"^[SABC]档_", td):
                existing_ids.update(f[:-4] for f in os.listdir(tdir) if f.endswith(".txt"))

    stats = {"index_rows": len(index), "index_file": index_name,
             "written": 0, "per_tier": {}, "missing_in_index": 0,
             "no_evidence": 0, "collide_existing": 0}
    for tier in tiers:
        dirname = TIER_DIRNAME[tier]
        src_dir = os.path.join(src, dirname)
        if not os.path.isdir(src_dir):
            print(f"[ingest] 源目录缺失，跳过 {tier} 档: {src_dir}")
            continue
        dest_dir = os.path.join(dest_root, dirname)
        os.makedirs(dest_dir, exist_ok=True)
        n = 0
        for fn in sorted(os.listdir(src_dir)):
            if not fn.endswith(".txt"):
                continue
            pid = fn[:-4]
            meta = index.get(pid)
            if meta is None:
                stats["missing_in_index"] += 1
                meta = {"arXiv ID": pid}   # 裸回退：档位来自目录，其余字段从正文推导
            if str(meta.get("赛题分档") or "").strip() != tier:
                print(f"[ingest] 分档不一致（xlsx={meta.get('赛题分档')} 目录={tier}）: {pid}")
            if not meta.get("证据句"):
                stats["no_evidence"] += 1
            if pid in existing_ids and not os.path.exists(os.path.join(dest_dir, fn)):
                stats["collide_existing"] += 1
                print(f"[ingest] 警告: {pid} 已存在于既有语料（重复入库风险）")
            with open(os.path.join(src_dir, fn), encoding="utf-8", errors="replace") as f:
                raw_body = f.read()
            if len(re.sub(r"\s+", " ", raw_body).strip()) < config.MIN_CONTENT_CHARS:
                print(f"[ingest] 跳转过短文件: {fn}")
                continue
            with open(os.path.join(dest_dir, fn), "w", encoding="utf-8", newline="\n") as f:
                f.write(convert_text(raw_body, meta, tier, batch))
            n += 1
        stats["per_tier"][tier] = n
        stats["written"] += n

    # 批次分档统计报告随档留存（溯源）
    for rep in glob.glob(os.path.join(src, "05_*.txt")):
        shutil.copyfile(rep, os.path.join(dest_root, os.path.basename(rep)))
    return stats


def main():
    ap = argparse.ArgumentParser(description="arXiv 全库批次入库（S/A 档 → 标准头块格式）")
    ap.add_argument("--src", required=True, help="批次目录（含 01_索引 xlsx 与档位文件夹）")
    ap.add_argument("--batch", default="arxiv2022", help="data/papers 下的批次目录名")
    ap.add_argument("--tiers", default="S,A", help="入库档位（默认 S,A；B/C 不入库）")
    ap.add_argument("--papers-dir", default=None, help="覆盖论文根目录（默认 data/papers）")
    args = ap.parse_args()
    tiers = tuple(t.strip().upper() for t in args.tiers.split(",") if t.strip())
    bad = [t for t in tiers if t not in TIER_DIRNAME]
    if bad:
        ap.error(f"不支持的档位 {bad}（可选 {'/'.join(TIER_DIRNAME)}）")

    stats = ingest(args.src, batch=args.batch, tiers=tiers, papers_dir=args.papers_dir)
    print(f"[ingest] 索引 {stats['index_file']} 共 {stats['index_rows']} 行")
    print(f"[ingest] 写入 {stats['written']} 篇："
          + "，".join(f"{t} 档 {c} 篇" for t, c in sorted(stats["per_tier"].items())))
    if stats["missing_in_index"]:
        print(f"[ingest] 索引缺行（裸回退转换）{stats['missing_in_index']} 篇")
    if stats["no_evidence"]:
        print(f"[ingest] 无证据句 {stats['no_evidence']} 篇（合法，头块省略该字段）")
    if stats["collide_existing"]:
        print(f"[ingest] 与既有语料同 ID {stats['collide_existing']} 篇，请人工核对！")
    print("[ingest] 完成。重建映射表：cd codes/timeline && python run_timeline.py --papers")


if __name__ == "__main__":
    main()
